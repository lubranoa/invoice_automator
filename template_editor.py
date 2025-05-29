from openpyxl import load_workbook

def main():
    filename = "excel_templates/Aus - No Program - Blank Invoice Template.xlsx"
    wb=load_workbook(filename=filename)
    print(wb.sheetnames)

if __name__=="__main__":
    main()